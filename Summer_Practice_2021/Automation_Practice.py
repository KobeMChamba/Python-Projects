import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    # Loads a xl file and returns a work book object
    wb = xl.load_workbook(filename)
    # Returns our first sheet, it is case sensitive
    sheet = wb['Sheet1']
    # we can get a cell using
    cell = sheet['a1']
    # or
    cell = sheet.cell(1, 1)
    # Use to find amount of rows
    # print(sheet.max_row)
    for row in range(2, sheet.max_row + 1):  # 2 3 4 because we want to ignore first row
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(filename)

process_workbook('transactions.xlsx')